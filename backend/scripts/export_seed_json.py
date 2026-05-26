#!/usr/bin/env python3
"""Export grant_funding_portals.xlsx and Opportunities.xlsx to committed JSON files."""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
BACKEND = Path(__file__).resolve().parents[1]
DATA_DIR = BACKEND / "data"

DEDICATED_SCRAPER_MAP = {
    "grants.gov": "grants_gov",
    "simpler.grants.gov": "grants_gov",
    "reporter.nih.gov": "nih_reporter",
    "nsf.gov": "nsf",
    "sbir.gov": "sbir",
    "ec.europa.eu": "eu_funding",
    "eufundingportal": "eu_funding",
    "gtr.ukri.org": "ukri_gtr",
    "iatistandard.org": "iati",
    "d-portal.org": "iati",
    "grantnav.threesixtygiving.org": "three60giving",
    "openalex.org": "openalex",
    "propublica.org": "propublica",
}

FUNDER_LOGO_DOMAINS = {
    "grants.gov": "/logos/grants-gov.svg",
    "simpler.grants.gov": "/logos/grants-gov.svg",
    "reporter.nih.gov": "/logos/nih.svg",
    "nsf.gov": "/logos/nsf.svg",
    "sbir.gov": "/logos/nsf.svg",
    "gatesfoundation.org": "/logos/gates-foundation.svg",
    "wellcome.org": "/logos/wellcome.svg",
    "fordfoundation.org": "/logos/ford-foundation.svg",
    "macfound.org": "/logos/macarthur.svg",
    "ec.europa.eu": "/logos/horizon-europe.svg",
    "cordis.europa.eu": "/logos/horizon-europe.svg",
    "erc.europa.eu": "/logos/horizon-europe.svg",
    "eic.ec.europa.eu": "/logos/horizon-europe.svg",
    "ukri.org": "/logos/ukri.svg",
    "gtr.ukri.org": "/logos/ukri.svg",
    "worldbank.org": "/logos/world-bank.svg",
    "theglobalfund.org": "/logos/global-fund.svg",
    "who.int": "/logos/who.svg",
    "unicef.org": "/logos/unicef.svg",
    "undp.org": "/logos/undp.svg",
    "usaid.gov": "/logos/usaid.svg",
    "edctp.org": "/logos/edctp.jpg",
    "elrha.org": "/logos/wellcome.svg",
    "ted.com": "/logos/ted.svg",
    "chanzuckerberg.com": "/logos/chan-zuckerberg.svg",
    "openphilanthropy.org": "/logos/open-philanthropy.svg",
}

HIGH_PRIORITY_NAMES = {
    "grants.gov (simpler grants)",
    "nih reporter",
    "nsf award search",
    "sbir.gov",
    "eu funding & tenders portal",
    "horizon europe / cordis",
    "ukri gateway to research",
    "ukri funding opportunities",
    "iati datastore / d-portal",
    "360giving grantnav",
    "openalex",
    "bill & melinda gates foundation",
    "wellcome trust",
    "grand challenges",
    "bill & melinda gates foundation – grand challenges explorations",
    "chan zuckerberg initiative",
    "wellcome leap",
}

API_ENDPOINT_MAP = {
    "grants.gov (simpler grants)": {
        "endpoint": "https://simpler.grants.gov/api/opportunities/search",
        "config": {
            "method": "post",
            "params": {"query": "health AI digital", "pagination": {"page_offset": 1, "page_size": 100}},
            "items_key": "data",
            "title_field": "opportunity_title",
            "desc_field": "summary_description",
            "url_field": "opportunity_id",
            "deadline_field": "close_date",
        },
    },
    "nih reporter": {
        "endpoint": "https://api.reporter.nih.gov/v2/projects/search",
        "config": {
            "method": "post",
            "items_key": "results",
            "title_field": "project_title",
            "desc_field": "abstract_text",
            "url_field": "opportunity_number",
            "deadline_field": "fiscal_year",
        },
    },
    "nsf award search": {
        "endpoint": "https://api.nsf.gov/services/v1/awards.json",
        "config": {
            "params": {"keyword": "artificial intelligence health digital", "dateStart": "01/01/2024"},
            "items_key": "response.award",
            "title_field": "title",
            "desc_field": "abstractText",
            "url_field": "id",
            "deadline_field": "expDate",
        },
    },
    "ukri gateway to research (gtr)": {
        "endpoint": "https://gtr.ukri.org/gtr/api/projects",
        "config": {
            "params": {"q": "AI health digital", "f": "pro.t", "s": 100},
            "items_key": "project",
            "title_field": "title",
            "desc_field": "abstractText",
            "url_field": "url",
        },
    },
}


def _get_logo_url(website: str) -> str | None:
    if not website:
        return None
    try:
        from urllib.parse import urlparse
        parsed = urlparse(website if website.startswith("http") else f"https://{website}")
        domain = parsed.netloc.lstrip("www.")
        for key, logo_path in FUNDER_LOGO_DOMAINS.items():
            if key in domain or domain in key:
                return logo_path
    except Exception:
        pass
    return None


def _get_scraper_type(name: str, website: str, has_api: str) -> str:
    website_lower = (website or "").lower()
    for key in DEDICATED_SCRAPER_MAP:
        if key in website_lower:
            return DEDICATED_SCRAPER_MAP[key]
    if has_api and (has_api.startswith("YES") or has_api.startswith("Partial")):
        return "ai_scraper"
    return "ai_scraper"


def export_sources(excel_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Grant & Funding Sources"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    sources = []
    for row in rows:
        if not any(row):
            continue
        category, name, website, has_api, *_rest = (row + (None,) * 8)[:8]
        scraping_notes = _rest[3] if len(_rest) > 3 else None
        data_available = _rest[2] if len(_rest) > 2 else None
        if not name or not website:
            continue
        name = str(name).strip()
        website = str(website).strip()
        has_api = str(has_api).strip() if has_api else "NO"
        category = str(category).strip() if category else "Other"
        scraping_notes = str(scraping_notes).strip() if scraping_notes else None
        data_available = str(data_available).strip() if data_available else None
        scraper_type = _get_scraper_type(name, website, has_api)
        is_high_priority = name.lower() in HIGH_PRIORITY_NAMES
        refresh_frequency = "daily" if is_high_priority else "weekly"
        scraper_config: dict = {}
        api_endpoint = None
        for key, api_cfg in API_ENDPOINT_MAP.items():
            if key in name.lower():
                api_endpoint = api_cfg["endpoint"]
                scraper_config = api_cfg.get("config", {})
                break
        if scraper_type == "ai_scraper" and not scraper_config:
            scraper_config = {"use_playwright": True, "crawl_depth": 0}
            if any(kw in website.lower() for kw in ["grants.gov", "foundation.org", "philanthropy", "fundsfor"]):
                scraper_config["crawl_depth"] = 1
        relevant_themes = []
        if data_available:
            d = data_available.lower()
            if any(kw in d for kw in ["health", "medical", "ai", "digital", "climate"]):
                relevant_themes = ["health", "digital", "AI"]
            if "development" in d or "lmic" in d or "international" in d:
                relevant_themes.append("global health")
        sources.append({
            "name": name,
            "category": category,
            "url": website,
            "source_type": scraper_type,
            "api_endpoint": api_endpoint,
            "scraper_config": scraper_config,
            "refresh_frequency": refresh_frequency,
            "is_high_priority": is_high_priority,
            "logo_url": _get_logo_url(website),
            "status": "active",
            "relevant_themes": list(set(relevant_themes)),
            "notes": scraping_notes,
        })
    return sources


def tier_to_score(tier: str, raw_score: float) -> tuple[float, str]:
    if tier in ("A★",):
        score = 82 + (raw_score - 4.1) * 40
    elif tier == "A":
        score = 68 + (raw_score - 3.6) * 28
    elif tier == "B":
        score = 54 + (raw_score - 3.2) * 35
    elif tier == "C":
        score = 40 + (raw_score - 2.8) * 35
    else:
        score = max(0, raw_score * 10)
    score = round(min(100.0, max(0.0, score)), 1)
    if score >= 80:
        priority = "high_priority"
    elif score >= 60:
        priority = "worth_reviewing"
    elif score >= 40:
        priority = "watchlist"
    else:
        priority = "low_fit"
    return score, priority


def parse_deadline(raw: str) -> date | None:
    if not raw or str(raw).lower() in ("tbd / rolling", "rolling", "ongoing (donor cultivation)", ""):
        return None
    raw = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw[:10], fmt[: len(raw[:10].split("-")[0]) + 6]).date()
        except ValueError:
            pass
    m = re.search(r"(\d{4}-\d{2}-\d{2})", raw)
    if m:
        try:
            return date.fromisoformat(m.group(1))
        except ValueError:
            pass
    return None


def export_opportunities(excel_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path)
    seen_titles: set[str] = set()
    all_opps: list[dict] = []
    for sheet_name in wb.sheetnames:
        if "Guide" in sheet_name or "Legend" in sheet_name:
            continue
        ws = wb[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                if row[0] == "Rank":
                    headers = [str(c).strip().replace("\n", "") if c else "" for c in row]
                continue
            if not row[0] or not row[1]:
                continue

            def col(name: str):
                try:
                    return row[headers.index(name)]
                except (ValueError, IndexError):
                    return None

            title = str(col("Opportunity") or "").strip()
            if not title or title in seen_titles:
                continue
            seen_titles.add(title)
            tier = str(col("Tier") or "D").strip()
            try:
                raw_score = float(col("SCORE") or 0)
            except (TypeError, ValueError):
                raw_score = 0.0
            fit_score, priority = tier_to_score(tier, raw_score)
            deadline = parse_deadline(str(col("Deadline") or ""))
            funding_str = str(col("Funding") or "")
            numbers = re.findall(r"[\d.]+", funding_str)
            currency = "USD"
            if "€" in funding_str or "EUR" in funding_str:
                currency = "EUR"
            elif "CHF" in funding_str:
                currency = "CHF"
            elif "£" in funding_str or "GBP" in funding_str:
                currency = "GBP"
            award_min = award_max = None
            if numbers:
                try:
                    vals = [float(n) for n in numbers[:2]]

                    def scale(v, s):
                        if "M" in s:
                            return v * 1_000_000
                        if "K" in s or "k" in s:
                            return v * 1_000
                        return v

                    award_min = scale(vals[0], funding_str)
                    award_max = scale(vals[-1], funding_str) if len(vals) > 1 else award_min
                except Exception:
                    pass
            notes = str(col("Notes") or "").strip()
            url = str(col("URL") or "").strip()
            funder = str(col("Funder") or "").strip()
            category = str(col("Category") or "").strip()
            rationale = f"[{tier}] Score: {raw_score:.2f}. {notes[:300]}" if notes else f"[{tier}] Score: {raw_score:.2f}"
            all_opps.append({
                "seed_key": title.lower().strip(),
                "title": title,
                "funder": funder,
                "program_name": category,
                "fit_score": fit_score,
                "priority": priority,
                "fit_rationale": rationale,
                "deadline": deadline.isoformat() if deadline else None,
                "award_min": award_min,
                "award_max": award_max,
                "currency": currency,
                "opportunity_url": url if url.startswith("http") else None,
                "notes": notes[:500] if notes else None,
                "status": "new",
                "thematic_areas": [],
                "keywords": [],
                "geography": [],
            })
    return all_opps


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--portals", default=str(ROOT / "grant_funding_portals.xlsx"))
    parser.add_argument("--opportunities", default=str(ROOT / "Opportunities.xlsx"))
    parser.add_argument("--out-dir", default=str(DATA_DIR))
    args = parser.parse_args()
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    portals_path = Path(args.portals)
    if portals_path.exists():
        sources = export_sources(portals_path)
        out = out_dir / "grant_funding_portals.json"
        out.write_text(json.dumps({"version": 1, "sources": sources}, indent=2))
        print(f"Wrote {len(sources)} sources → {out}")
    else:
        print(f"Skip sources: {portals_path} not found")

    opps_path = Path(args.opportunities)
    if opps_path.exists():
        opps = export_opportunities(opps_path)
        out = out_dir / "grant_opportunities_seed.json"
        out.write_text(json.dumps({"version": 1, "opportunities": opps}, indent=2))
        print(f"Wrote {len(opps)} opportunities → {out}")
    else:
        print(f"Skip opportunities: {opps_path} not found")


if __name__ == "__main__":
    main()
