#!/usr/bin/env python3
"""
Seed script: import the 133 LiGHT opportunities from Opportunities.xlsx
into the database as pre-scored opportunities.

Usage:
    DATABASE_URL=postgresql://light:light@localhost:5432/light_grants python scripts/seed_opportunities.py
    Or just: python scripts/seed_opportunities.py (uses config.yaml)

The Excel scoring maps to fit_score as follows:
    Tier A★ (score ≥4.1) → 85–100  (high_priority)
    Tier A  (score ≥3.6) → 70–84   (worth_reviewing)
    Tier B  (score ≥3.2) → 55–69   (watchlist)
    Tier C  (score ≥2.8) → 40–54   (watchlist)
    Tier D  (<2.8)       → 0–39    (low_fit)
"""
import os
import sys
import uuid
from pathlib import Path
from datetime import date, datetime
import re

import openpyxl

# Allow running from repo root
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

# ── Database setup ─────────────────────────────────────────────────────────────
import sqlalchemy as sa
from sqlalchemy.orm import Session

def get_engine():
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        # Try config.yaml
        import yaml
        config_path = Path(__file__).parent.parent / "config.yaml"
        with open(config_path) as f:
            cfg = yaml.safe_load(f)
        db_url = cfg["database"]["url"]
    return sa.create_engine(db_url)


def tier_to_score(tier: str, raw_score: float) -> tuple[float, str]:
    """Convert Excel tier + raw score to 0-100 fit_score and priority label."""
    if tier in ("A★",):
        score = 82 + (raw_score - 4.1) * 40  # 82–100
    elif tier == "A":
        score = 68 + (raw_score - 3.6) * 28   # 68–82
    elif tier == "B":
        score = 54 + (raw_score - 3.2) * 35   # 54–68
    elif tier == "C":
        score = 40 + (raw_score - 2.8) * 35   # 40–54
    else:
        score = max(0, raw_score * 10)          # <40

    score = round(min(100.0, max(0.0, score)), 1)

    if score >= 80:
        priority = "high_priority"
    elif score >= 60:
        priority = "worth_reviewing"
    elif score >= 40:
        priority = "watchlist"
    else:
        priority = "low_fit"

    return score, priority


def parse_deadline(raw: str) -> date | None:
    """Parse deadline strings from the Excel sheet."""
    if not raw or str(raw).lower() in ("tbd / rolling", "rolling", "ongoing (donor cultivation)", ""):
        return None
    raw = str(raw).strip()
    # Try ISO date first
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw[:10], fmt[:len(raw[:10].split("-")[0])+6]).date()
        except ValueError:
            pass
    # Extract year-month-day with regex
    m = re.search(r"(\d{4}-\d{2}-\d{2})", raw)
    if m:
        try:
            return date.fromisoformat(m.group(1))
        except ValueError:
            pass
    return None


def load_opportunities_xlsx(path: str) -> list[dict]:
    """Load all opportunities from the Excel file across all relevant sheets."""
    wb = openpyxl.load_workbook(path)
    seen_titles = set()
    all_opps = []

    for sheet_name in wb.sheetnames:
        if "Guide" in sheet_name or "Legend" in sheet_name:
            continue
        ws = wb[sheet_name]
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if headers is None:
                # Find header row: look for 'Rank' in first column
                if row[0] == "Rank":
                    headers = [str(c).strip().replace("\n", "") if c else "" for c in row]
                continue

            if not row[0] or not row[1]:
                continue

            # Map columns
            def col(name: str):
                try:
                    idx = headers.index(name)
                    return row[idx]
                except (ValueError, IndexError):
                    return None

            title = str(col("Opportunity") or "").strip()
            if not title or title in seen_titles:
                continue
            seen_titles.add(title)

            tier = str(col("Tier") or "D").strip()
            try:
                raw_score = float(col("SCORE") or 0)
            except (TypeError, ValueError):
                raw_score = 0.0

            fit_score, priority = tier_to_score(tier, raw_score)

            deadline_raw = col("Deadline")
            deadline = parse_deadline(str(deadline_raw) if deadline_raw else "")

            funding_str = str(col("Funding") or "")
            # Parse award amount from strings like "$500K–$3M", "€8–15M", "CHF 200K–1M"
            numbers = re.findall(r"[\d.]+", funding_str)
            currency = "USD"
            if "€" in funding_str or "EUR" in funding_str:
                currency = "EUR"
            elif "CHF" in funding_str:
                currency = "CHF"
            elif "£" in funding_str or "GBP" in funding_str:
                currency = "GBP"

            award_min = None
            award_max = None
            if numbers:
                try:
                    vals = [float(n) for n in numbers[:2]]
                    # Convert K/M suffixes
                    def scale(v, s):
                        if "M" in s: return v * 1_000_000
                        if "K" in s or "k" in s: return v * 1_000
                        return v
                    award_min = scale(vals[0], funding_str)
                    award_max = scale(vals[-1], funding_str) if len(vals) > 1 else award_min
                except Exception:
                    pass

            notes = str(col("Notes") or "").strip()
            url = str(col("URL") or "").strip()
            funder = str(col("Funder") or "").strip()
            category = str(col("Category") or "").strip()
            stage = str(col("Stage") or "").strip()

            # Build fit rationale from notes
            rationale = f"[{tier}] Score: {raw_score:.2f}. {notes[:300]}" if notes else f"[{tier}] Score: {raw_score:.2f}"

            all_opps.append({
                "id": str(uuid.uuid4()),
                "title": title,
                "funder": funder,
                "program_name": category,
                "fit_score": fit_score,
                "priority": priority,
                "fit_rationale": rationale,
                "deadline": deadline,
                "award_min": award_min,
                "award_max": award_max,
                "currency": currency,
                "opportunity_url": url if url.startswith("http") else None,
                "notes": notes[:500] if notes else None,
                "status": "needs_review" if fit_score >= 40 else "new",
                "thematic_areas": [],
                "keywords": [],
                "geography": [],
                "date_discovered": datetime.utcnow(),
                "date_updated": datetime.utcnow(),
            })

    return all_opps


def seed(xlsx_path: str):
    engine = get_engine()
    opps = load_opportunities_xlsx(xlsx_path)
    print(f"Loaded {len(opps)} opportunities from Excel")

    with Session(engine) as db:
        inserted = 0
        for opp_data in opps:
            existing = db.execute(
                sa.text("SELECT id FROM opportunities WHERE title = :title"),
                {"title": opp_data["title"]}
            ).fetchone()
            if existing:
                continue

            db.execute(
                sa.text("""
                    INSERT INTO opportunities
                    (id, title, funder, program_name, fit_score, priority, fit_rationale,
                     deadline, award_min, award_max, currency, opportunity_url, notes,
                     status, thematic_areas, keywords, geography, date_discovered, date_updated)
                    VALUES
                    (:id, :title, :funder, :program_name, :fit_score, :priority, :fit_rationale,
                     :deadline, :award_min, :award_max, :currency, :opportunity_url, :notes,
                     :status, :thematic_areas::jsonb, :keywords::jsonb, :geography::jsonb,
                     :date_discovered, :date_updated)
                """),
                {
                    **opp_data,
                    "thematic_areas": "[]",
                    "keywords": "[]",
                    "geography": "[]",
                    "deadline": str(opp_data["deadline"]) if opp_data["deadline"] else None,
                }
            )
            inserted += 1

        db.commit()
        print(f"Inserted {inserted} new opportunities")


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "Opportunities.xlsx"
    if not Path(xlsx_path).exists():
        print(f"Error: {xlsx_path} not found")
        print("Usage: python scripts/seed_opportunities.py path/to/Opportunities.xlsx")
        sys.exit(1)
    seed(xlsx_path)
